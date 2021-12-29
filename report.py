from openpyxl import load_workbook
import os

COLOR_ERROR = '#8b0000'
COLOR_OK = '#006400'

class Report:
    """Report class representation.
    """

    def __init__(self, name, result, details):
        """Initialization function for the Report class.

        Args:
            name (string): name of the Report
            result (string): result of the evaluation, can either be None, "OK" or "ERROR".
            details (string): details of the evaluation. Will contain found lines from the report.
        """
        self.name = name
        self.result = result
        self.details = details


def evaluate(v1, v2, operator):
    """Evaluation function for two values and an operator.

    Args:
        v1 (obj): object one
        v2 (obj): object two
        operator (string): operator that is being used

    Returns:
        bool: result of the equation v1 operator v2
    """
    if operator == '==':
        return v1 == v2
    elif operator == '<=':
        return v1 <= v2
    elif operator == '>=':
        return v1 >= v2
    elif operator == '!=':
        return v1 != v2
    elif operator == '>':
        return v1 > v2
    elif operator == '<':
        return v1 < v2

def report_thread(queue, filename, working_directory, rule):
    """Function to represent a report evaluation.
    Will load the report using the provided filename and working_directory and 
    evaluate it according to the provided rule.

    Args:
        queue (Queue): queue for the results to be put into
        filename (string): filename
        working_directory (string): path to the directory of the reports
        rule (Rule): used rule for evaluation
    """
    wb = load_workbook(os.path.join(working_directory, filename))

    lines = list()
    result = Report(filename, '', '')
    
    # iterate the rows
    found = False
    for line in wb.active.iter_rows():
        for c in line:
            if c.column == rule.col or rule.col < 0:
                if rule.type == 'BGCOLOR':
                    # check if the color is indexed for legacy colors
                    col = ''
                    if c.fill.bgColor.type == 'indexed':
                        col = c.fill.fgColor.rgb
                    else:
                        col = c.fill.bgColor.rgb
                    
                    if col == rule.compare:
                        found = True
                elif rule.type == 'FONTCOLOR':
                    if c.font != None and c.font.color != None:
                        if c.font.color.rgb == rule.compare:
                            found = True
                elif rule.type == 'VALUE' and (type(c.value) == int or type(c.value) == float):
                    if evaluate(c.value, rule.compare, rule.operator):
                        found = True
                        
        # search criteria was found
        if found:
            new_line = ''
            for c in line:
                new_line += str(c.value)
                if c != line[-1]:
                    new_line += ', '
            new_line += '\n'
            lines.append(new_line)
            found = False
        
    if len(lines) == 0:
        result.result = 'OK'
    else:
        result.result = 'ERROR'
        # save error lines
        for line in lines:
            result.details += line
            result.details += '\n'

    lines.clear()
    queue.put(result)